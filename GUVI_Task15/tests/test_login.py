from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from GUVI_Task15.pages.login_page import LoginPage
import openpyxl
from datetime import datetime
from selenium.webdriver.common.by import By

def test_login(driver):

    workbook = openpyxl.load_workbook("GUVI_Task15/testdata/login_data.xlsx")

    sheet = workbook.active

    for row in range(2,sheet.max_row+1):

        username = sheet.cell(row=row, column=2).value
        password = sheet.cell(row=row,column=3).value

        login_page = LoginPage(driver)

        login_page.open_url("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")

        WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH,"//input[@name='username']")))

        print(f"Executing Row: {row}")
        print("username:", username)
        print("password:", password)

        login_page.login(username,password)

        current_date = datetime.now().strftime("%d-%m-%y")
        current_time = datetime.now().strftime("%H:%M:%S")

        sheet.cell(row=row,column=4).value = current_date
        sheet.cell(row=row,column=5).value = current_time

        tester_name = sheet.cell(row=row,column=6).value

        if "dashboard" in driver.current_url.lower():
            sheet.cell(row=row,column=7).value = "Test Passed"
        else:
            sheet.cell(row=row,column=7).value = "Test Failed"


    workbook.save("GUVI_Task15/testdata/login_data.xlsx")